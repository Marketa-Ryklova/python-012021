import pandas
vykazy = pandas.read_csv("vykazy.csv", index_col = 0)
print(vykazy.head())
hodiny_projekty = vykazy.groupby('project')['hours'].sum()
print(hodiny_projekty)
hodiny_projekty.to_excel("hodiny_projekty.xlsx")
#uložení do xlxs

excel_data = pandas.read_excel('hodiny_projekty.xlsx', index_col=0)
print(excel_data)
#načtení z excelu

from openpyxl import Workbook

wb = Workbook()
ws1 = wb.active
ws1.title = "rozvrh"
cislo_radku = 1
cislo_sloupce = 2

bunka = ws1.cell(cislo_radku, cislo_sloupce)
bunka.value = "Test"

wb.save(filename="rozvrh_hodin.xlsx")


from openpyxl import Workbook
from openpyxl.styles import PatternFill

wb = Workbook()
ws1 = wb.active
ws1.title = "rozvrh"

rozvrh_hodin = [
  ["Anglický jazyk", "Přírodopis", "Dějepis", "Matematika", "Oběd", "Tělesná výchova", "Tělesná výchova", ],
  ["Občanská výchova", "Hudební výchova", "Matematika", "Oběd", "Výtvarná výchova", "Dějepis", ],
  ["Matematika", "Chemie", "Přírodopis", "Fyzika", "Oběd", "Zeměpis", ],
  ["Fyzika", "Anglický jazyk", "Matematika", "Český jazyk", "Dějepis", "Oběd", ],
  ["Český jazyk", "Zeměpis", "Český jazyk", "Výtvarná výchova", "Oběd", ]
]
radek = 1
for den in rozvrh_hodin:
  sloupec = 1
  for predmet in den:
    bunka = ws1.cell(cislo_radku, cislo_sloupce)
    bunka.value = rozvrh_hodin[3][4]
    sloupec += 1
  radek += 1
#vypsala jsem předmět Dějepis, který se koná ve čtvrtek před obědem
wb.save(filename="rozvrh_hodin.xlsx")

